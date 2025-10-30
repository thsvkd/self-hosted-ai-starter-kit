from openpyxl import load_workbook
from datetime import datetime
import json
import sys
from pathlib import Path

input_file = "/data/shared/주차 할인권 구입 신청서.xlsx"

try:
    wb = load_workbook(input_file)
    ws = wb.active

    # determine new date string (실행 월로)
    today = datetime.now()
    year = today.year
    month = today.month
    day = 24
    new_date_str = f"{year}년 {month:02d}월 {day:02d}일"

    # 병합셀 upper-left 셀(D14) 내용만 변경하면 병합상태, 포맷 유지
    ws["D14"].value = new_date_str

    output_file = f"/data/shared/주차 할인권 구입 신청서 {month + 1:02d}월.xlsx"
    # 저장
    wb.save(output_file)

    # Print JSON to stdout so the caller (api_server) can parse it
    # out = {"output_file": str(Path(output_file).resolve())}
    print(str(Path(output_file).resolve()), flush=True, end="")
    sys.exit(0)

except Exception as e:
    # Print structured error to stderr and exit non-zero so api_server can report it
    try:
        err = {"error": str(e)}
        print(json.dumps(err), file=sys.stderr, flush=True)
    except Exception:
        print(str(e), file=sys.stderr, flush=True)
    sys.exit(2)
